from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from .models import (
    Kategori, Eslesme, Bildirim, UserProfile, 
    BlogKategori, BlogYazisi, BlogYorum, SSS
)
from .forms import KategoriForm, EslesmeForm, CustomUserCreationForm
from .scraping_utils import akakce_fiyat_cek, site_fiyat_cek, fiyat_farki_hesapla
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
from django.http import FileResponse, JsonResponse
from django_ratelimit.decorators import ratelimit
from decouple import config

def export_to_excel_dark(df, excel_path):
    """Excel dosyasını koyu tema ile formatlar"""
    columns = list(df.columns)
    akakce_link_kolonu = "Akakçe Linki"
    site_link_kolonu = "Sitenizdeki Link"
    new_columns = [c for c in columns if c not in [akakce_link_kolonu, site_link_kolonu]] + [akakce_link_kolonu, site_link_kolonu]
    df = df.reindex(columns=new_columns)
    df.to_excel(excel_path, index=False)
    wb = load_workbook(excel_path)
    ws = wb.active

    # Stil tanımlamaları
    header_fill = PatternFill(start_color="FF1e3a8a", end_color="FF1e3a8a", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF", name="Segoe UI", size=12)
    zebra1 = PatternFill(start_color="FF3b82f6", end_color="FF3b82f6", fill_type="solid")
    zebra2 = PatternFill(start_color="FF1e40af", end_color="FF1e40af", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    # Header stilini uygula
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Zebra pattern uygula
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = zebra1 if i % 2 == 0 else zebra2
        for cell in row:
            cell.fill = fill
            cell.font = Font(name="Segoe UI", size=11, color="FFEEEEFF")
            cell.alignment = Alignment(vertical="center")

    # Sütun genişliklerini ayarla
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value)
                if val is None: 
                    val = ""
                if len(val) > max_length:
                    max_length = len(val)
            except:
                pass
        new_width = min(max_length + 2, 60)
        ws.column_dimensions[column].width = new_width

    # Link sütunlarını gizle
    for idx, cell in enumerate(ws[1], 1):
        if cell.value in [akakce_link_kolonu, site_link_kolonu]:
            ws.column_dimensions[cell.column_letter].hidden = True

    wb.save(excel_path)

def home_view(request):
    """Ana sayfa görünümü"""
    # Son blog yazıları
    son_yazilar = BlogYazisi.objects.filter(yayinlandi=True)[:3]
    
    context = {
        'son_yazilar': son_yazilar,
    }
    return render(request, "home.html", context)

def register(request):
    """Kullanıcı kayıt görünümü"""
    if request.user.is_authenticated:
        return redirect('profile')
    
    if request.method == "POST":
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            # UserProfile'ı manuel olarak oluştur
            UserProfile.objects.get_or_create(user=user)
            messages.success(request, "Kayıt başarılı! Giriş yapabilirsiniz.")
            return redirect('login')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = CustomUserCreationForm()
    
    return render(request, "register.html", {"form": form})

def login_view(request):
    """Kullanıcı giriş görünümü"""
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect("fiyat_cek")
        else:
            messages.error(request, "Kullanıcı adı veya şifre hatalı.")
    
    return render(request, "login.html")

def logout_view(request):
    """Kullanıcı çıkış görünümü"""
    logout(request)
    return redirect("login")

@login_required
def profile(request):
    """Kullanıcı profil görünümü"""
    # UserProfile'ı kontrol et ve oluştur
    user_profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    kategoriler = Kategori.objects.filter(user=request.user)
    kategori_id = request.GET.get("kategori")
    aktif_kategori = None
    eslesmeler = []
    eslesme_form = EslesmeForm()
    kategori_form = KategoriForm()

    if kategori_id:
        aktif_kategori = get_object_or_404(Kategori, id=kategori_id, user=request.user)
        eslesmeler = Eslesme.objects.filter(kategori=aktif_kategori)

    if request.method == "POST":
        if "ekle_kategori" in request.POST:
            kategori_form = KategoriForm(request.POST)
            if kategori_form.is_valid():
                kategori = kategori_form.save(commit=False)
                kategori.user = request.user
                kategori.save()
                bildirim_ekle(request.user, f"Yeni kategori eklendi: {kategori.isim}", notification_type="success")
                messages.success(request, "Kategori eklendi.")
                return redirect(f"{request.path}?kategori={kategori.id}")
        
        elif "sil_kategori" in request.POST:
            sil_id = request.POST.get("sil_kategori_id")
            sil_kat = Kategori.objects.filter(id=sil_id, user=request.user)
            if sil_kat.exists():
                kategori_ismi = sil_kat[0].isim
                sil_kat.delete()
                bildirim_ekle(request.user, f"Kategori silindi: {kategori_ismi}", notification_type="warning")
                messages.success(request, "Kategori silindi.")
                kalan = Kategori.objects.filter(user=request.user).first()
                return redirect(f"{request.path}?kategori={kalan.id if kalan else ''}")
        
        elif "eslesme_ekle" in request.POST and kategori_id:
            eslesme_form = EslesmeForm(request.POST)
            if eslesme_form.is_valid():
                eslesme = eslesme_form.save(commit=False)
                eslesme.kategori = aktif_kategori
                eslesme.save()
                bildirim_ekle(request.user, f"Yeni eşleşme eklendi: {eslesme.akakce_link}", notification_type="success")
                messages.success(request, "Eşleşme eklendi.")
                return redirect(f"{request.path}?kategori={aktif_kategori.id}")
        
        elif "eslesme_sil" in request.POST:
            sil_id = request.POST.get("eslesme_sil_id")
            eslesme = Eslesme.objects.filter(id=sil_id, kategori__user=request.user)
            if eslesme.exists():
                kategori_id = eslesme[0].kategori.id
                bildirim_ekle(request.user, f"Eşleşme silindi: {eslesme[0].akakce_link}", notification_type="warning")
                eslesme.delete()
                messages.success(request, "Eşleşme silindi.")
                return redirect(f"{request.path}?kategori={kategori_id}")

    context = {
        "kategoriler": kategoriler,
        "aktif_kategori": aktif_kategori,
        "eslesmeler": eslesmeler,
        "eslesme_form": eslesme_form,
        "kategori_form": kategori_form,
        "user": request.user,
        "messages": messages.get_messages(request),
    }
    return render(request, "profile.html", context)

@ratelimit(key='user', rate='10/m', block=True)
@login_required
def fiyat_cek(request):
    """Fiyat çekme ana görünümü"""
    # UserProfile'ı kontrol et ve oluştur
    user_profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    kategoriler = Kategori.objects.filter(user=request.user)
    kategori_id = request.GET.get("kategori")
    aktif_kategori = None
    eslesmeler = []
    eslesme_form = EslesmeForm()
    kategori_form = KategoriForm()
    fiyat_sonuclari = []

    if kategori_id:
        aktif_kategori = get_object_or_404(Kategori, id=kategori_id, user=request.user)
        eslesmeler = Eslesme.objects.filter(kategori=aktif_kategori)

    if request.method == "POST":
        if "ekle_kategori" in request.POST:
            kategori_form = KategoriForm(request.POST)
            if kategori_form.is_valid():
                kategori = kategori_form.save(commit=False)
                kategori.user = request.user
                kategori.save()
                bildirim_ekle(request.user, f"Yeni kategori eklendi: {kategori.isim}", notification_type="success")
                messages.success(request, "Kategori eklendi.")
                return redirect(f"{request.path}?kategori={kategori.id}")
        
        elif "sil_kategori" in request.POST:
            sil_id = request.POST.get("sil_kategori_id")
            sil_kat = Kategori.objects.filter(id=sil_id, user=request.user)
            if sil_kat.exists():
                kategori_ismi = sil_kat[0].isim
                sil_kat.delete()
                bildirim_ekle(request.user, f"Kategori silindi: {kategori_ismi}", notification_type="warning")
                messages.success(request, "Kategori silindi.")
                kalan = Kategori.objects.filter(user=request.user).first()
                return redirect(f"{request.path}?kategori={kalan.id if kalan else ''}")
        
        elif "eslesme_ekle" in request.POST and kategori_id:
            eslesme_form = EslesmeForm(request.POST)
            if eslesme_form.is_valid():
                eslesme = eslesme_form.save(commit=False)
                eslesme.kategori = aktif_kategori
                eslesme.save()
                bildirim_ekle(request.user, f"Yeni eşleşme eklendi: {eslesme.akakce_link}", notification_type="success")
                messages.success(request, "Eşleşme eklendi.")
                return redirect(f"{request.path}?kategori={aktif_kategori.id}")
        
        elif "eslesme_sil" in request.POST:
            sil_id = request.POST.get("eslesme_sil_id")
            eslesme = Eslesme.objects.filter(id=sil_id, kategori__user=request.user)
            if eslesme.exists():
                kategori_id = eslesme[0].kategori.id
                bildirim_ekle(request.user, f"Eşleşme silindi: {eslesme[0].akakce_link}", notification_type="warning")
                eslesme.delete()
                messages.success(request, "Eşleşme silindi.")
                return redirect(f"{request.path}?kategori={kategori_id}")
        
        elif "fiyatlari_cek" in request.POST and kategori_id:
            kategori = get_object_or_404(Kategori, id=kategori_id, user=request.user)
            eslesmeler = Eslesme.objects.filter(kategori=kategori)
            
            data = []
            for eslesme in eslesmeler:
                # Akakçe'den fiyat çek
                akakce_result = akakce_fiyat_cek(eslesme.akakce_link)
                
                # Kendi siteden fiyat çek
                site_fiyat = site_fiyat_cek(eslesme.site_link) if eslesme.site_link else "Site linki yok"
                
                # Fiyat farkını hesapla
                fiyat_farki = fiyat_farki_hesapla(akakce_result['fiyat'], site_fiyat)
                
                row_data = {
                    "urun_adi": akakce_result['urun_adi'],
                    "akakce_fiyati": akakce_result['fiyat'],
                    "akakce_magaza": akakce_result['magaza'],
                    "akakce_linki": eslesme.akakce_link,
                    "site_linki": eslesme.site_link or "",
                    "site_fiyati": site_fiyat,
                    "fiyat_farki": fiyat_farki,
                }
                data.append(row_data)
                fiyat_sonuclari.append(row_data)
            
            # Excel dosyası oluştur ve indir
            if "excel_indir" in request.POST:
                df = pd.DataFrame(data)
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    export_to_excel_dark(df, tmp.name)
                    response = FileResponse(
                        open(tmp.name, "rb"), 
                        as_attachment=True, 
                        filename=f"akakce_buybox_{kategori.isim}.xlsx"
                    )
                    bildirim_ekle(request.user, f"{kategori.isim} kategorisi için fiyatlar Excel olarak indirildi.", notification_type="info")
                    return response
            else:
                # Web'de göster
                bildirim_ekle(request.user, f"{kategori.isim} kategorisi için fiyatlar çekildi.", notification_type="success")
                messages.success(request, f"{len(data)} ürün için fiyat bilgileri çekildi.")
        
        return redirect(f"{request.path}?kategori={kategori_id or ''}")

    context = {
        "kategoriler": kategoriler,
        "aktif_kategori": aktif_kategori,
        "eslesmeler": eslesmeler,
        "eslesme_form": eslesme_form,
        "kategori_form": kategori_form,
        "fiyat_sonuclari": fiyat_sonuclari,
        "user": request.user,
        "messages": messages.get_messages(request),
    }
    return render(request, "fiyat_cek.html", context)

@login_required
def settings_view(request):
    """Ayarlar görünümü"""
    # UserProfile'ı kontrol et ve oluştur
    user_profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    kategoriler = Kategori.objects.filter(user=request.user)
    default_category_id = getattr(user_profile, "default_category_id", None)
    theme = request.session.get("theme", "light")

    if request.method == "POST":
        # E-posta güncelleme
        email = request.POST.get("email")
        if email and email != request.user.email:
            request.user.email = email
            request.user.save()
            messages.success(request, "E-posta güncellendi.")
        
        # Varsayılan kategori güncelleme
        default_category_id = request.POST.get("default_category")
        if default_category_id and default_category_id.isdigit():
            user_profile.default_category_id = int(default_category_id)
            user_profile.save()
            messages.success(request, "Varsayılan kategori güncellendi.")
        
        # Tema güncelleme
        theme = request.POST.get("theme", "light")
        request.session["theme"] = theme
        messages.success(request, "Tema tercihi kaydedildi.")

    context = {
        "user": request.user,
        "kategoriler": kategoriler,
        "default_category_id": default_category_id,
        "theme": theme,
    }
    return render(request, "ayarlar.html", context)

def bildirim_ekle(user, mesaj, notification_type="info"):
    """Kullanıcıya bildirim ekler"""
    Bildirim.objects.create(user=user, mesaj=mesaj, notification_type=notification_type)

@login_required
def bildirim_merkezi(request):
    """Bildirim merkezi görünümü"""
    bildirimler = request.user.bildirimler.order_by('-olusturma_zamani')
    
    if request.method == "POST":
        # Tüm bildirimleri okundu olarak işaretle
        request.user.bildirimler.filter(okundu=False).update(okundu=True)
        messages.success(request, "Tüm bildirimler okundu olarak işaretlendi.")
    
    return render(request, "bildirimler.html", {"bildirimler": bildirimler})

# Blog Views
def blog_liste(request):
    """Blog yazıları listesi"""
    yazilar = BlogYazisi.objects.filter(yayinlandi=True)
    kategoriler = BlogKategori.objects.all()
    
    # Arama
    q = request.GET.get('q')
    if q:
        yazilar = yazilar.filter(
            Q(baslik__icontains=q) | 
            Q(ozet__icontains=q) | 
            Q(icerik__icontains=q)
        )
    
    # Sayfalama
    paginator = Paginator(yazilar, 6)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'kategoriler': kategoriler,
        'q': q,
    }
    return render(request, 'blog/liste.html', context)

def blog_kategori(request, kategori_slug):
    """Blog kategori sayfası"""
    kategori = get_object_or_404(BlogKategori, slug=kategori_slug)
    yazilar = BlogYazisi.objects.filter(kategori=kategori, yayinlandi=True)
    
    # Sayfalama
    paginator = Paginator(yazilar, 6)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'kategori': kategori,
        'page_obj': page_obj,
    }
    return render(request, 'blog/kategori.html', context)

def blog_detay(request, slug):
    """Blog yazısı detay sayfası"""
    yazi = get_object_or_404(BlogYazisi, slug=slug, yayinlandi=True)
    
    # Görüntülenme sayısını artır
    yazi.goruntulenme_sayisi += 1
    yazi.save(update_fields=['goruntulenme_sayisi'])
    
    # İlgili yazılar
    ilgili_yazilar = BlogYazisi.objects.filter(
        kategori=yazi.kategori, 
        yayinlandi=True
    ).exclude(id=yazi.id)[:3]
    
    # Yorumlar
    yorumlar = yazi.yorumlar.filter(onaylandi=True)
    
    context = {
        'yazi': yazi,
        'ilgili_yazilar': ilgili_yazilar,
        'yorumlar': yorumlar,
    }
    return render(request, 'blog/detay.html', context)

def sss_view(request):
    """Sık Sorulan Sorular sayfası"""
    sss_listesi = SSS.objects.filter(aktif=True)
    
    context = {
        'sss_listesi': sss_listesi,
    }
    return render(request, 'sss.html', context)

